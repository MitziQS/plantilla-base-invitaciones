from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
DARK="3D2E1F"; GOLD="B08D57"; CREAM="F5EFE6"; WHITE="FDFAF6"; LGOLD="D4AF77"

def hdr(cell, txt, bg=DARK, fg=CREAM, sz=11):
    cell.value=txt; cell.font=Font(name="Arial",bold=True,color=fg,size=sz)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def bdr():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)

def dc(cell,val="",bold=False,color="3D2E1F",bg=None,center=False):
    cell.value=val; cell.font=Font(name="Arial",bold=bold,color=color,size=10)
    cell.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=True)
    if bg: cell.fill=PatternFill("solid",start_color=bg)
    cell.border=bdr()

# SHEET 1 — CLIENTES
ws1=wb.active; ws1.title="Clientes"
ws1.sheet_view.showGridLines=False; ws1.freeze_panes="A3"
ws1.row_dimensions[1].height=38; ws1.row_dimensions[2].height=28
ws1.merge_cells("A1:N1")
c=ws1["A1"]; c.value="TRACKER DE CLIENTES — INVITACIONES DIGITALES"
c.font=Font(name="Arial",bold=True,color=CREAM,size=13)
c.fill=PatternFill("solid",start_color=DARK)
c.alignment=Alignment(horizontal="center",vertical="center")

cols=[("A","#",8),("B","Cliente",22),("C","Tipo evento",15),("D","Nombre(s)",22),
      ("E","Fecha evento",13),("F","Fecha entrega",13),("G","Estado",14),
      ("H","URL Invitación",30),("I","URL Sheet RSVPs",30),("J","Confirmados",12),
      ("K","Precio $",12),("L","Pagado",11),("M","Notas",25),("N","WhatsApp",16)]

for col,title,width in cols:
    hdr(ws1[col+"2"],title,bg=GOLD,fg=WHITE,sz=10)
    ws1.column_dimensions[col].width=width

estados={"En proceso":("FFF9C4","5F4F00"),"Entregado":("C8E6C9","1B5E20"),
          "Pagado":("BBDEFB","0D47A1"),"Pendiente":("FFCCBC","BF360C")}

rows=[
  [1,"Sofía Ramírez","Boda","Sofía & Mateo","14/02/2026","10/01/2026","Entregado",
   "https://plantilla-base-invitaciones.netlify.app/sofia-mateo.html","https://docs.google.com/...",12,2500,"Sí","Demo",""],
  [2,"Laura Gómez","XV Años","Isabella","20/03/2026","01/02/2026","En proceso",
   "","","",1800,"50% anticipo","Pendiente fotos","+52 55 9876 5432"],
  [3,"Carlos Mendoza","Cumpleaños","Rodrigo","05/04/2026","15/03/2026","Pendiente",
   "","","",1500,"No","","+52 55 5555 0000"],
]

for ri,row in enumerate(rows,start=3):
    ws1.row_dimensions[ri].height=22
    bg="FDFAF6" if ri%2==1 else "F0EAE0"
    for ci,(col,_,_) in enumerate(cols):
        cell=ws1[col+str(ri)]; v=row[ci]
        if col=="G":
            sb,sf=estados.get(v,("FFFFFF","000000")); dc(cell,v,bold=True,color=sf,bg=sb,center=True)
        elif col=="A": dc(cell,v,bold=True,center=True,bg=GOLD,color=WHITE)
        elif col in("H","I"): dc(cell,v,color="1565C0")
        elif col in("J","K","L"): dc(cell,v,center=True)
        else: dc(cell,v,bg=bg)

for ri in range(6,55):
    ws1.row_dimensions[ri].height=22
    bg="FDFAF6" if ri%2==1 else "F0EAE0"
    for col,_,_ in cols:
        cell=ws1[col+str(ri)]
        cell.fill=PatternFill("solid",start_color=bg); cell.border=bdr()
        cell.alignment=Alignment(vertical="center")

# SHEET 2 — DASHBOARD
ws2=wb.create_sheet("Dashboard")
ws2.sheet_view.showGridLines=False
ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=18
ws2.row_dimensions[1].height=38
ws2.merge_cells("A1:B1")
c=ws2["A1"]; c.value="DASHBOARD — RESUMEN DEL NEGOCIO"
c.font=Font(name="Arial",bold=True,color=CREAM,size=13)
c.fill=PatternFill("solid",start_color=DARK)
c.alignment=Alignment(horizontal="center",vertical="center")

metrics=[
  ("MÉTRICAS GENERALES",None,True),
  ("Total clientes","=COUNTA(Clientes!B3:B54)",False),
  ("Entregados","=COUNTIF(Clientes!G3:G54,\"Entregado\")",False),
  ("En proceso","=COUNTIF(Clientes!G3:G54,\"En proceso\")",False),
  ("Pendientes","=COUNTIF(Clientes!G3:G54,\"Pendiente\")",False),
  ("",None,False),
  ("INGRESOS",None,True),
  ("Total facturado","=SUM(Clientes!K3:K54)",False),
  ("Total cobrado","=SUMIF(Clientes!L3:L54,\"Sí\",Clientes!K3:K54)",False),
  ("Por cobrar","=B8-B9",False),
  ("",None,False),
  ("POR TIPO",None,True),
  ("Bodas","=COUNTIF(Clientes!C3:C54,\"Boda\")",False),
  ("XV Años","=COUNTIF(Clientes!C3:C54,\"XV Años\")",False),
  ("Cumpleaños","=COUNTIF(Clientes!C3:C54,\"Cumpleaños\")",False),
  ("Bautizos","=COUNTIF(Clientes!C3:C54,\"Bautizo\")",False),
  ("Graduaciones","=COUNTIF(Clientes!C3:C54,\"Graduación\")",False),
]

for ri,(label,formula,is_sec) in enumerate(metrics,start=2):
    ws2.row_dimensions[ri].height=24
    if is_sec:
        ws2.merge_cells(f"A{ri}:B{ri}"); hdr(ws2["A"+str(ri)],label,bg=GOLD,fg=WHITE,sz=10)
    elif label=="":
        pass
    else:
        a=ws2["A"+str(ri)]; b=ws2["B"+str(ri)]
        a.value=label; a.font=Font(name="Arial",size=10,color=DARK)
        a.fill=PatternFill("solid",start_color="FDFAF6"); a.border=bdr()
        a.alignment=Alignment(vertical="center",indent=1)
        b.value=formula; b.font=Font(name="Arial",bold=True,size=11,color=DARK)
        b.fill=PatternFill("solid",start_color=CREAM); b.border=bdr()
        b.alignment=Alignment(horizontal="center",vertical="center")
        if "facturado" in label.lower() or "cobrado" in label.lower() or "cobrar" in label.lower():
            b.number_format='$#,##0'

# SHEET 3 — GUIA
ws3=wb.create_sheet("Guia de uso")
ws3.sheet_view.showGridLines=False
ws3.column_dimensions["A"].width=28; ws3.column_dimensions["B"].width=55
ws3.row_dimensions[1].height=38
ws3.merge_cells("A1:B1")
c=ws3["A1"]; c.value="GUÍA DE USO — FLUJO POR CLIENTE"
c.font=Font(name="Arial",bold=True,color=CREAM,size=13)
c.fill=PatternFill("solid",start_color=DARK)
c.alignment=Alignment(horizontal="center",vertical="center")

guide=[
  ("FLUJO DE TRABAJO",None,True),
  ("1. Cliente nuevo","Agrega una fila en Clientes con nombre, tipo de evento y fecha"),
  ("2. Crea su Google Sheet","sheets.new → Extensiones → Apps Script → pega el código → Implementar"),
  ("3. Genera la invitación","Panel admin → selecciona tema → llena datos → pega URL Script → Generar"),
  ("4. Sube a Netlify","Arrastra el HTML a Netlify → copia la URL pública"),
  ("5. Actualiza el tracker","Pega URL invitación (col H) y URL Sheet RSVPs (col I)"),
  ("6. Comparte con cliente","Envía la URL por WhatsApp. Activa el QR para impresión física"),
  ("7. Monitorea RSVPs","Abre la URL del Sheet del cliente (col I) para ver quién confirma"),
  ("8. Cierra el cliente","Actualiza estado a Pagado y marca en col L"),
  ("",None,False),
  ("PRECIOS SUGERIDOS",None,True),
  ("Paquete Básico — $1,500","9 slides + RSVP Google Sheets + QR imprimible"),
  ("Paquete Estándar — $2,000","Básico + foto o video embed + música"),
  ("Paquete Premium — $3,000","Estándar + dominio personalizado + soporte 30 días"),
  ("Ajuste post-entrega — $300","Por cambio adicional después de entrega"),
  ("",None,False),
  ("ESTADOS",None,True),
  ("Pendiente","Cliente interesado, trabajo no iniciado"),
  ("En proceso","Invitación en creación o ajuste"),
  ("Entregado","Invitación activa y enviada al cliente"),
  ("Pagado","Pago completo recibido"),
]

for ri,(a_val,b_val,*rest) in enumerate(guide,start=2):
    ws3.row_dimensions[ri].height=22
    is_sec=rest[0] if rest else False
    if is_sec:
        ws3.merge_cells(f"A{ri}:B{ri}"); hdr(ws3["A"+str(ri)],a_val,bg=GOLD,fg=WHITE,sz=10)
    elif a_val=="":
        pass
    else:
        a=ws3["A"+str(ri)]; b=ws3["B"+str(ri)]
        a.value=a_val; a.font=Font(name="Arial",bold=True,size=10,color=DARK)
        a.fill=PatternFill("solid",start_color="F0EAE0"); a.border=bdr()
        a.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
        b.value=b_val; b.font=Font(name="Arial",size=10,color="3D2E1F")
        b.fill=PatternFill("solid",start_color="FDFAF6"); b.border=bdr()
        b.alignment=Alignment(vertical="center",wrap_text=True,indent=1)

wb.save("/sessions/upbeat-lucid-brahmagupta/mnt/outputs/tracker-clientes.xlsx")
print("DONE")
